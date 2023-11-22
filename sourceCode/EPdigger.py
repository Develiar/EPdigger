import numpy as np
import os
from Bio import SeqIO
import math
import re
from PySide2.QtWidgets import QApplication, QMessageBox, QFileDialog
from PySide2.QtUiTools import QUiLoader
from PySide2.QtCore import QFile, Signal, QObject
from threading import Thread
import os.path
import openpyxl
from PySide2.QtGui import QIcon



class SignalStore(QObject):
    
    progress_update = Signal(int)
    
    finished = Signal(list)
    
    error_window = Signal(int)






so = SignalStore()


class AboutWidget():
    def __init__(self):
        
        ui_file = QFile('aboutWin.ui')  
        ui_file.open(QFile.ReadOnly)
        ui_file.close()

        
        
        
        self.ui = QUiLoader().load(ui_file)


class MainWindow():

    def __init__(self):
        
        so.progress_update.connect(self.setProgress)
        so.finished.connect(self.show_message)
        so.error_window.connect(self.window_error)

        
        ui_file = QFile('mainWin.ui')  
        ui_file.open(QFile.ReadOnly)
        ui_file.close()

        
        
        
        self.ui = QUiLoader().load(ui_file)

        
        self.filePaths_ref = []
        self.filePaths_alfa = []
        self.filePaths_rsa = []

        
        self.ui.choice_Button_fa.clicked.connect(self.get_path_ref)

        
        self.ui.choice_Button_alfa.clicked.connect(self.get_path_alfa)

        
        self.ui.choice_Button_rsa.clicked.connect(self.get_path_rsa)

        
        self.ui.progressBar.setRange(0, 10)

        
        self.ui.runButton.clicked.connect(self.runp)

        
        self.ui.actionAbout.triggered.connect(self.open_AboutWidget)

    def open_AboutWidget(self):
        
        self.aboutw = AboutWidget()
        
        self.aboutw.ui.show()

    
    def setProgress(self, value):
        self.ui.progressBar.setValue(value)

    
    def window_error(self, number):
        
        message = ""
        ew = QMessageBox()
        ew.setWindowTitle("Error")
        ew.setIcon(QMessageBox.Warning)
        if number == 1:
            message = "Error in all full-length sequence file"
        elif number == 2:
            message = "Error in RSA file"
        elif number == 3:
            message = "Please check that these files are correct"
        elif number == 4:
            message = "All full-length sequence file have sequences containing unknown or non-standard amino acids"
        elif number == 5:
            message = "Error in All full-length sequences have only one sequence"
        elif number == 6:
            message = "Please close pv_ASA.xlsx"
        elif number == 7:
            message = "Please close w_value.xlsx"
        elif number == 8:
            message = "Please close result.xlsx"
        else:
            message = "Error in reference sequence file"
        ew.setText(message)
        ew.exec_()

    
    def show_message(self, message):
        
        result_msg = ""
        msg = QMessageBox()
        msg.setWindowTitle("Result")
        for i in range(len(message[0])):
            result_msg += (message[0][i] + '\n' + message[1][i]) + '\n'
        msg.setText(result_msg)
        msg.setIcon(QMessageBox.Information)
        
        msg.exec_()

    def get_path_ref(self):
        path1, _ = QFileDialog.getOpenFileName(self.ui, "Select file", "", "Fasta Files (*.Fasta)")
        self.filePaths_ref.append(path1)
        filename = os.path.basename(path1)
        self.ui.label_5.setText(filename)


    def get_path_alfa(self):
        path2, _ = QFileDialog.getOpenFileName(self.ui, "Select file", "", "Fasta Files (*.Fasta)")
        self.filePaths_alfa.append(path2)
        filename = os.path.basename(path2)
        self.ui.label_6.setText(filename)


    def get_path_rsa(self):
        path3, _ = QFileDialog.getOpenFileName(self.ui, "Select file", "", "Text Files (*.txt)")
        self.filePaths_rsa.append(path3)
        filename = os.path.basename(path3)
        self.ui.label_7.setText(filename)


    def runp(self):
        
        ckLengt = self.ui.spinBox.value()

        

        
        def d20(seq):
            
            aa_dict = {
                'A': [1.001, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'C': [0, 1.002, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'D': [0, 0, 1.003, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'E': [0, 0, 0, 1.004, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'F': [0, 0, 0, 0, 1.005, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'G': [0, 0, 0, 0, 0, 1.006, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'H': [0, 0, 0, 0, 0, 0, 1.007, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'I': [0, 0, 0, 0, 0, 0, 0, 1.008, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'K': [0, 0, 0, 0, 0, 0, 0, 0, 1.009, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'L': [0, 0, 0, 0, 0, 0, 0, 0, 0, 1.010, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'M': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.011, 0, 0, 0, 0, 0, 0, 0, 0, 0],
                'N': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.012, 0, 0, 0, 0, 0, 0, 0, 0],
                'P': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.013, 0, 0, 0, 0, 0, 0, 0],
                'Q': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.014, 0, 0, 0, 0, 0, 0],
                'R': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.015, 0, 0, 0, 0, 0],
                'S': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.016, 0, 0, 0, 0],
                'T': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.017, 0, 0, 0],
                'V': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.018, 0, 0],
                'W': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.019, 0],
                'Y': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1.020]}

            
            result = [aa_dict[aa] for aa in seq]

            
            a = 0
            b = 0
            lst1 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            lst2 = []
            while a < len(result):
                lst1 = list(np.add(lst1, result[b]))
                a += 1
                b += 1

                lst2.append(lst1)  

            
            a1 = 0
            b1 = 0
            lst3 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            while a1 < len(lst2):
                lst3 = list(np.add(lst3, lst2[b1]))
                a1 += 1
                b1 += 1

            return lst3

        
        def gswnPv(seq):
            lt = len(seq)
            n = ckLengt  
            prl = []  
            for a in range(lt):  
                
                left = a  
                right = a + (ckLengt - 1)  
                if right > lt - 1:
                    
                    break
                elif right < (ckLengt - 1):
                    
                    break
                seq1 = seq[left:right + 1]
                seq2 = d20(seq1)  
                ui = [(x / n) ** 2 for x in seq2]
                pr = math.sqrt(sum(ui))
                prl.append(pr)
            return prl

        
        def re1pv(pv):
            pv1 = []
            for i in pv:
                p = 1 / i
                pv1.append(p)
            return pv1

        
        def nor(pv1):
            c = max(pv1)
            d = min(pv1)
            b = c - d
            ynl = []
            if b != 0:
                for x in pv1:
                    a = x - d
                    yn = a / b * 10
                    ynl.append(yn)
            else:
                ynl = [1 for i in pv1]

            return ynl

        def gswnAsa(seq):
            lt = len(seq)
            asal = []  

            for a in range(lt):  
                
                left = a  
                right = a + (ckLengt - 1)  
                if right > lt - 1:
                    
                    break
                elif right < (ckLengt - 1):
                    
                    break
                seq1 = seq[left:right + 1]
                x = 0
                for i in seq1:
                    x += i
                asa = x / ckLengt
                asal.append(asa)

            return asal

        
        def is_single_sequence(file_path):
            with open(file_path, "r") as handle:
                records = list(SeqIO.parse(handle, "fasta"))
            return len(records) == 1

        def calculate():
            self.ui.runButton.setEnabled(False)
            so.progress_update.emit(0)
            errNum = 3
            try:
                
                filename1 = ''
                refilename = self.filePaths_ref[-1]
                if is_single_sequence(refilename):
                    filename1 = self.filePaths_alfa[-1]
                else:
                    errNum = 0
                    a1 > b1  

                so.progress_update.emit(1)

                if is_single_sequence(filename1):
                    errNum = 5
                    a1 > b1  
                else:
                    for record in SeqIO.parse(filename1, "fasta"):
                        if "X" in record.seq or "Z" in record.seq or "U" in record.seq or "B" in record.seq:
                            errNum = 4
                            a1 > b1  

                so.progress_update.emit(2)

                try:
                    pRl = [gswnPv(seq_record.seq) for seq_record in SeqIO.parse(filename1, "fasta")]
                except Exception:
                    errNum = 1
                so.progress_update.emit(3)

                
                transposed = [[row[i] for row in pRl] for i in range(len(pRl[0]))]

                
                pvlst = [len(set(i)) for i in transposed]

                
                pv1 = re1pv(pvlst)  
                pvn = nor(pv1)  
                

                
                data_pvASA = []

                line1 = [((str(i)) + ' to ' + (str(i + ckLengt - 1))) for i in range(1, len(pRl[0]) + 1)]
                line1.insert(0, ' ')

                data_pvASA.append(line1)
                
                linePv = pvlst[:]
                linePv.insert(0, 'PV')
                data_pvASA.append(linePv)

                so.progress_update.emit(4)

                
                try:
                    with open(self.filePaths_rsa[-1], 'r') as f:
                        data = f.read()

                    numbers = []
                    for line in data.split('\n'):
                        if not line.startswith('>'):
                            numbers += re.findall(r'\d+', line)

                    num = [int(x) for x in numbers]

                    asa = gswnAsa(num)
                    asa1 = asa[:]  
                    asan = nor(asa)
                    
                except Exception:
                    errNum = 2

                asa.insert(0, 'ASA')
                data_pvASA.append(asa)

                try:
                    filename = "./result/pv_ASA.xlsx"

                    if os.path.isfile(filename):
                        
                        i = 1
                        while os.path.isfile(f'{filename}.{i}.xlsx'):
                            i += 1
                        os.rename(filename, f'{filename}.{i}.xlsx')

                    
                    workbook = openpyxl.Workbook()

                    
                    worksheet = workbook.active

                    
                    for row in data_pvASA:
                        worksheet.append(row)

                    
                    workbook.save('./result/pv_ASA.xlsx')
                    so.progress_update.emit(5)
                except Exception:
                    errNum = 6
                    a1 > b1

                
                wlst = list(np.add(pvn, asan))
                

                
                w_index = list(enumerate(wlst))
                w_sorted = sorted(w_index, key=lambda x: x[1], reverse=True)

                

                
                seqlen = 0
                w_select = []  
                for record in SeqIO.parse(filename1, "fasta"):
                    
                    seqlen = len(record.seq)
                    break
                if seqlen > 1000:
                    w_select = w_sorted[:100]
                elif 500 <= seqlen <= 1000:
                    w_select = w_sorted[:75]
                else:
                    w_select = w_sorted[:50]
                

                
                wtrue = [[x[0] + 1, x[1]] for x in w_select]
                wtrue.insert(0, ['Starting position of the peptides', 'W value'])

                try:
                    filename = "./result/w_value.xlsx"

                    if os.path.isfile(filename):
                        
                        i = 1
                        while os.path.isfile(f'{filename}.{i}.xlsx'):
                            i += 1
                        os.rename(filename, f'{filename}.{i}.xlsx')

                    
                    workbook = openpyxl.Workbook()

                    
                    worksheet = workbook.active

                    
                    for row in wtrue:
                        worksheet.append(row)

                    
                    workbook.save('./result/w_value.xlsx')

                    so.progress_update.emit(6)

                except Exception:
                    errNum = 7
                    a1 > b1

                winl = [w[0] for w in w_select]  
                
                win = sorted(winl, reverse=False)
                aa_position = [i + 1 for i in win]  

                

                

                
                def group_list(lst):
                    groups = []
                    group = [lst[0]]
                    for i in range(1, len(lst)):
                        if lst[i] - lst[i - 1] >= ckLengt:
                            groups.append(group)
                            group = [lst[i]]
                        else:
                            group.append(lst[i])
                    groups.append(group)
                    return groups

                aa_group = group_list(aa_position)
                
                h = [lst[-1] - lst[0] + ckLengt for lst in aa_group]  
                

                
                result = []  
                for tpl in aa_group:
                    result.append(list(range(tpl[0], tpl[-1] + 1)))

                pv1gl = [list(pv1[i - 1] for i in gr) for gr in result]  
                asagl = [list(asa1[i - 1] for i in gr) for gr in result]  

                
                pv1g_mean = [np.mean(x) for x in pv1gl]  
                asag_mean = [np.mean(x) for x in asagl]  

                pv1gn = nor(pv1g_mean)
                
                asagn = nor(asag_mean)
                
                hn = nor(h)
                
                so.progress_update.emit(7)

                alst = []
                blst = []
                clst = []
                slst = []
                for i in range(len(hn)):
                    a = math.sqrt(pv1gn[i] ** 2 + hn[i] ** 2 + pv1gn[i] * hn[i])
                    alst.append(a)

                    b = math.sqrt(hn[i] ** 2 + asagn[i] ** 2 + hn[i] * asagn[i])
                    blst.append(b)

                    c = math.sqrt(asagn[i] ** 2 + pv1gn[i] ** 2 + asagn[i] * pv1gn[i])
                    clst.append(c)

                    s = (a + b + c) / 2
                    slst.append(s)

                so.progress_update.emit(8)
                abc_area = []  
                for k in range(len(slst)):
                    area = math.sqrt(slst[k] * (slst[k] - alst[k]) * (slst[k] - blst[k]) * (slst[k] - clst[k]))
                    abc_area.append(area)
                

                area_sorted = sorted(enumerate(abc_area), key=lambda x: x[1], reverse=True)
                area_top_50percent = area_sorted[:int(len(area_sorted) * 0.5)]

                

                abcindex = [lst[0] for lst in area_top_50percent]

                

                abc_50index = [result[x] for x in abcindex]
                
                aa_start_end = []

                for aa in abc_50index:
                    if len(aa) > 1:
                        aa[-1] = aa[-1] + ckLengt - 1
                        aa_start_end.append([aa[0], aa[-1]])
                    else:
                        aa_start_end.append([aa[0], aa[-1] + ckLengt - 1])
                
                so.progress_update.emit(9)

                
                data_result = []

                ref = []
                refilename = self.filePaths_ref[-1]

                for seq_record in SeqIO.parse(refilename, "fasta"):
                    ref = seq_record.seq

                aresult = [ref[aa_position[0] - 1:aa_position[1]] for aa_position in aa_start_end]
                aa_result = [str(seq) for seq in aresult]

                aa_start_end_result = [','.join(map(str, i)) for i in aa_start_end]
                
                data_result.append(aa_start_end_result)
                data_result.append(aa_result)

                try:
                    filename = "./result/result.xlsx"
                    if os.path.isfile(filename):
                        
                        i = 1
                        while os.path.isfile(f'{filename}.{i}.xlsx'):
                            i += 1
                        os.rename(filename, f'{filename}.{i}.xlsx')

                    
                    workbook = openpyxl.Workbook()

                    
                    worksheet = workbook.active

                    
                    for row in data_result:
                        worksheet.append(row)

                    
                    workbook.save('./result/result.xlsx')

                    
                    so.progress_update.emit(10)
                except Exception:
                    errNum = 8
                    a1 > b1
                self.ui.runButton.setEnabled(True)
                
                so.finished.emit(data_result)
                

            except Exception:
                so.error_window.emit(errNum)
                self.ui.runButton.setEnabled(True)



        
        worker = Thread(target=calculate)
        
        worker.start()
        
        


if __name__ == '__main__':
    
    app = QApplication([])
    
    app.setWindowIcon(QIcon('logo.png'))
    
    window = MainWindow()
    
    window.ui.show()
    
    app.exec_()
